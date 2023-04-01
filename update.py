import os
import requests
import openpyxl
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, numbers
from forex_python.converter import CurrencyRates

def convert_usd_to_eur(price_usd):
    return currency_converter.convert('USD', 'EUR', price_usd)

def get_skin_price_eur(skin_name, session):
    search_url = f"https://steamcommunity.com/market/search?l=german&category_730_ItemSet%5B%5D=any&q={skin_name}"
    response = session.get(search_url)
    soup = BeautifulSoup(response.text, 'html.parser')
    items = soup.find_all('a', class_='market_listing_row_link')

    for item in items:
        item_name = item.find('span', class_='market_listing_item_name').text
        if skin_name.lower() in item_name.lower():
            price_usd_element = item.find('span', class_='sale_price') or item.find('span', class_='normal_price')
            if price_usd_element:
                price_usd_text = price_usd_element.text.strip()
                price_usd = float(price_usd_text.replace("USD", "").strip()[1:])
                price_eur = convert_usd_to_eur(price_usd)
                return price_eur
    return None

# Load the existing Excel file
excel_file_path = os.path.join(os.getcwd(), "investments.xlsx")
workbook = load_workbook(excel_file_path)
worksheet = workbook.active

currency_converter = CurrencyRates()

# Create a session
session = requests.Session()

# Iterate through the rows, starting from the second row (skipping the header)
total_profit = 0
for row in range(2, worksheet.max_row + 1):
    skin_name = worksheet.cell(row=row, column=1).value

    if skin_name:
        price_eur = get_skin_price_eur(skin_name, session)
        print(price_eur)
        if price_eur is not None:
            price_eur = round(price_eur, 2)
            cell = worksheet.cell(row=row, column=4, value=price_eur)
            cell.number_format = '#,##0.00\ "€";[Red]\-#,##0.00\ "€"'

            purchase_price = worksheet.cell(row=row, column=3).value
            profit = price_eur - purchase_price
            profit_cell = worksheet.cell(row=row, column=5, value=profit)
            profit_cell.number_format = '#,##0.00\ "€";[Red]\-#,##0.00\ "€"'
            amount = worksheet.cell(row=row, column=2).value
            total_profit_per_item = profit * amount
            total_profit += total_profit_per_item
            total_profit_cell = worksheet.cell(row=row, column=6, value=total_profit_per_item)
            total_profit_cell.number_format = '#,##0.00\ "€";[Red]\-#,##0.00\ "€"'


        else:
            worksheet.cell(row=row, column=4, value="Not Found")

all_total_profit_cell = worksheet.cell(row=2, column=7, value=total_profit)
all_total_profit_cell.number_format = '#,##0.00\ "€";[Red]\-#,##0.00\ "€"'

# Save the updated Excel file
workbook.save(excel_file_path)
