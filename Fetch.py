import os
import requests
import openpyxl
import urllib.parse
import re
import time
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, numbers
from datetime import datetime, timedelta

#Global Variable Declaration
global language
global update_hours
global file_name
global currency
global currency_code
global currency_token


try:
    with open('settings.txt', 'r') as f:
        for line in f:
            if line.startswith('language'):
                language = line.split('=')[1].strip()
            elif line.startswith('file_name'):
                file_name = line.split('=')[1].strip()
            elif line.startswith('currency'):
                currency = line.split('=')[1].strip()
            elif line.startswith('update_hours'):
                update_hours = int(line.split('=')[1].strip())
except FileNotFoundError:
    f = open("settings.txt", "w")
    f.write("language=english\nfile_name=Investments.xlsx\ncurrency=EUR\nupdate_hours=24")
    f.close()
    print("Didnt find settings.txt. Created new Settings with default Config")
    time.sleep(1)
    with open('settings.txt', 'r') as f:
        for line in f:
            if line.startswith('language'):
                language = line.split('=')[1].strip()
            elif line.startswith('file_name'):
                file_name = line.split('=')[1].strip()
            elif line.startswith('currency'):
                currency = line.split('=')[1].strip()
            elif line.startswith('update_hours'):
                update_hours = int(line.split('=')[1].strip())

if currency.casefold() == "usd":
    currency_code = 1
    currency_token = "$"
elif currency.casefold() == "gbp":
    currency_code = 2
    currency_token = "£"
elif currency.casefold() == "eur":
    currency_code = 3
    currency_token = "€"

def get_skin_price(market_hash_name, session):
    price_overview_url = f"https://steamcommunity.com/market/priceoverview/?appid=730&currency={currency_code}&market_hash_name={market_hash_name}"
    response = session.get(price_overview_url)
    data = response.json()
    if data and data['success']:
        lowest_price = data['lowest_price']
        cleaned_price = re.sub(r"[^0-9.,]", "", lowest_price).replace(",", ".")
        price = float(cleaned_price)
        return price
    return None

def get_market_hash_name(skin_name, session):
    search_url = f"https://steamcommunity.com/market/search?appid=730&q={skin_name}&l={language}"
    response = session.get(search_url)
    if response.status_code == 429:
        return None, "too_many_requests"
    soup = BeautifulSoup(response.text, 'html.parser')
    search_results_row = soup.find('div', {'id': 'result_0'})
    if search_results_row:
        market_hash_name = search_results_row['data-hash-name']
        return market_hash_name, None
    return None, "unavailable"

# Load the existing Excel file
excel_file_path = os.path.join(os.getcwd(), file_name)
workbook = load_workbook(excel_file_path)
worksheet = workbook.active


# Check if the File is open right now
try:
    workbook.save(excel_file_path)
except PermissionError:
    print("Excel File is still open. Closing Excel File")
    os.system("taskkill /f /im excel.exe")

# Create a separate worksheet for storing market hash names
if "MarketHashNames" not in workbook:
    market_hash_names_worksheet = workbook.create_sheet("MarketHashNames")
else:
    market_hash_names_worksheet = workbook["MarketHashNames"]

# Create a session
session = requests.Session()

# Iterate through the rows, starting from the second row (skipping the header)
total_profit = 0
for row in range(2, worksheet.max_row + 1):
    skin_name = worksheet.cell(row=row, column=1).value
    if skin_name is not None:
        market_hash_name_cell = market_hash_names_worksheet.cell(row=row, column=2)
        market_hash_name = market_hash_name_cell.value
        # Check if the saved skin name in the MarketHashNames worksheet matches the skin name in the main table
        saved_skin_name_cell = market_hash_names_worksheet.cell(row=row, column=1)
        saved_skin_name = saved_skin_name_cell.value
        if saved_skin_name != skin_name:
            market_hash_name = None
            market_hash_name_cell.value = None
        last_updated_cell = worksheet.cell(row=row, column=7)
        last_updated = last_updated_cell.value
        update_price = True
        if last_updated is not None:
            time_since_last_update = datetime.now() - last_updated
            if time_since_last_update.total_seconds() < timedelta(hours=update_hours).total_seconds():
                update_price = False
        if update_price:
            if market_hash_name is None:
                parse_skin_name = urllib.parse.quote(skin_name)
                market_hash_name, error = get_market_hash_name(parse_skin_name, session)
                time.sleep(2)  # Add delay after get_market_hash_name
                if market_hash_name is not None:
                    # Set market_hash_name and skin_name in the worksheet
                    market_hash_name_cell.value = market_hash_name
                    saved_skin_name_cell.value = skin_name
                    workbook.save(excel_file_path)
                elif error == "too_many_requests":
                    print(f"Error 429: Too many requests. Try again later for Item: {skin_name}")
                    continue
                elif error == "unavailable":
                    print("Couldn't find Item: " + skin_name + ". Please check your spelling.")
            if market_hash_name is not None:
                market_hash_name = urllib.parse.quote(market_hash_name)
                price = get_skin_price(market_hash_name, session)
                time.sleep(3)  # Add delay after get_skin_price
                if price is not None:
                    cell = worksheet.cell(row=row, column=4, value=price)
                    cell.number_format = f'#,##0.00\ "{currency_token}";[Red]\-#,##0.00\ "{currency_token}"'
                    
                    # Set formulas for profit per item and total profit per item
                    profit_cell = worksheet.cell(row=row, column=5)
                    profit_cell.value = f"=D{row}-C{row}"
                    profit_cell.number_format = f'#,##0.00\ "{currency_token}";[Red]\-#,##0.00\ "{currency_token}"'
                    
                    total_profit_cell = worksheet.cell(row=row, column=6)
                    total_profit_cell.value = f"=E{row}*B{row}"
                    total_profit_cell.number_format = f'#,##0.00\ "{currency_token}";[Red]\-#,##0.00\ "{currency_token}"'

                    print("Successfully got Price for " + skin_name)
                    last_updated_cell.value = datetime.now()
                    # Save the updated Excel file
                    workbook.save(excel_file_path)
                else:
                    print(f"Couldn't get Price for {skin_name}, please try again later")
            else:
                print("Couldn't get Market Hash Name, please try again later")
        else:
            print(f"Skipping update for Item: {skin_name} as it was updated within the last {update_hours} hours")

# Save the updated Excel file
workbook.save(excel_file_path)
input("Press Enter  to close the program...")